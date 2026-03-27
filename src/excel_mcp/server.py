import logging
import mimetypes
import os
import secrets
import shutil
import json
import hmac
import hashlib
import time
from functools import wraps
from pathlib import Path
from typing import Any, List, Dict, Optional
from urllib.parse import quote, unquote

from mcp.server.fastmcp import FastMCP
from mcp.types import ToolAnnotations
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request
from starlette.responses import FileResponse, HTMLResponse, JSONResponse, Response

# Import exceptions
from excel_mcp.exceptions import (
    ValidationError,
    WorkbookError,
    SheetError,
    DataError,
    FormattingError,
    CalculationError,
    PivotError,
    ChartError
)

# Import from excel_mcp package with consistent _impl suffixes
from excel_mcp.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
    validate_range_in_sheet_operation as validate_range_impl
)
from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.workbook import get_workbook_info
from excel_mcp.data import write_data
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.tables import create_excel_table as create_table_impl
from excel_mcp.sheet import (
    copy_sheet,
    delete_sheet,
    rename_sheet,
    merge_range,
    unmerge_range,
    get_merged_ranges,
    insert_row,
    insert_cols,
    delete_rows,
    delete_cols,
)

# Get project root directory path for log file path.
# When using the stdio transmission method,
# relative paths may cause log files to fail to create
# due to the client's running location and permission issues,
# resulting in the program not being able to run.
# Thus using os.path.join(ROOT_DIR, "excel-mcp.log") instead.

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LOG_FILE = os.path.join(ROOT_DIR, "excel-mcp.log")

# Initialize EXCEL_FILES_PATH variable without assigning a value
EXCEL_FILES_PATH = None

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        # Referring to https://github.com/modelcontextprotocol/python-sdk/issues/409#issuecomment-2816831318
        # The stdio mode server MUST NOT write anything to its stdout that is not a valid MCP message.
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")
# Initialize FastMCP server
mcp = FastMCP(
    "excel-mcp",
    host=os.environ.get("FASTMCP_HOST", "0.0.0.0"),
    port=int(os.environ.get("FASTMCP_PORT", "8017")),
    instructions="Excel MCP Server for manipulating Excel files"
)

API_KEY_ENV_VAR = "EXCEL_MCP_API_KEY"
API_KEY_HEADER_ENV_VAR = "EXCEL_MCP_API_KEY_HEADER"
OUTPUT_DIR_PRIMARY_ENV_VAR = "DOC_OUTPUT_DIR"
OUTPUT_DIR_COMPAT_ENV_VAR = "MCP_OUTPUT_DIR"
DOWNLOAD_BASE_URL_PRIMARY_ENV_VAR = "DOC_DOWNLOAD_BASE_URL"
DOWNLOAD_BASE_URL_COMPAT_ENV_VAR = "MCP_DOWNLOAD_BASE_URL"
DOWNLOAD_SIGNING_SECRET_PRIMARY_ENV_VAR = "DOC_DOWNLOAD_SIGNING_SECRET"
DOWNLOAD_SIGNING_SECRET_COMPAT_ENV_VAR = "MCP_DOWNLOAD_SIGNING_SECRET"
DOWNLOAD_URL_TTL_PRIMARY_ENV_VAR = "DOC_DOWNLOAD_URL_TTL_SECONDS"
DOWNLOAD_URL_TTL_COMPAT_ENV_VAR = "MCP_DOWNLOAD_URL_TTL_SECONDS"

FINISHER_HINT_LINE_1 = "Finisher hint: when the user request appears complete, call"
FINISHER_HINT_TEXT = (
    f"{FINISHER_HINT_LINE_1} save_document_mcp_word-mcp so the user receives a download link."
)


def append_finisher_hint_to_text(response_text: str) -> str:
    """Append the finisher hint to plain text or JSON tool responses."""
    text = (response_text or "").rstrip()
    if FINISHER_HINT_LINE_1 in text:
        return response_text

    try:
        parsed = json.loads(text)
    except (TypeError, ValueError, json.JSONDecodeError):
        if not text:
            return FINISHER_HINT_TEXT
        return f"{text} {FINISHER_HINT_TEXT}"

    if isinstance(parsed, dict):
        message = parsed.get("message")
        if isinstance(message, str):
            parsed["message"] = append_finisher_hint_to_text(message)
        else:
            parsed["finisher_hint"] = FINISHER_HINT_TEXT
        return json.dumps(parsed, indent=2, default=str)

    return f"{text}\n\n{FINISHER_HINT_TEXT}"


def append_finisher_hint_to_tool_output(func):
    """Wrap a tool function so every string response includes the finisher hint."""

    @wraps(func)
    def wrapper(*args: Any, **kwargs: Any):
        response = func(*args, **kwargs)
        if isinstance(response, str):
            return append_finisher_hint_to_text(response)
        return response

    return wrapper


class APIKeyMiddleware(BaseHTTPMiddleware):
    """Simple API key middleware for HTTP transports."""

    def __init__(
        self,
        app: Any,
        api_key: str,
        header_name: str = "x-api-key",
        exempt_paths: Optional[List[str]] = None,
    ):
        super().__init__(app)
        self.api_key = api_key
        self.header_name = header_name.lower()
        self.exempt_paths = exempt_paths or []

    async def dispatch(self, request: Request, call_next: Any) -> Response:
        request_path = request.url.path
        if request.method.upper() == "OPTIONS":
            return await call_next(request)

        if request_path.startswith("/files/"):
            signed_status = evaluate_signed_download_request(request)
            if signed_status == "valid":
                return await call_next(request)
            if signed_status in {"expired", "invalid"}:
                return _signed_link_error_response(request, signed_status)

        for exempt in self.exempt_paths:
            if request_path == exempt or request_path.startswith(f"{exempt.rstrip('/')}/"):
                return await call_next(request)

        provided_key = request.headers.get(self.header_name)
        if not provided_key:
            return JSONResponse(
                {"error": f"Missing API key header: {self.header_name}"},
                status_code=401,
            )

        if not secrets.compare_digest(provided_key, self.api_key):
            return JSONResponse({"error": "Invalid API key."}, status_code=403)

        return await call_next(request)

def _clean_user_path(raw_path: Optional[str]) -> str:
    """Trim whitespace and wrapping quotes from user-provided paths."""
    value = (raw_path or "").strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in ("'", '"'):
        value = value[1:-1].strip()
    return value

def _is_basename_only(path_value: str) -> bool:
    path_obj = Path(path_value)
    return not path_obj.is_absolute() and path_obj.parent == Path(".")

def resolve_output_dir(default_if_missing: Optional[str] = None) -> Optional[str]:
    """
    Resolve output directory with compatibility fallbacks.

    Priority:
      1) DOC_OUTPUT_DIR
      2) EXCEL_FILES_PATH
      3) MCP_OUTPUT_DIR
      4) default_if_missing
    """
    configured = (
        os.environ.get(OUTPUT_DIR_PRIMARY_ENV_VAR)
        or os.environ.get("EXCEL_FILES_PATH")
        or os.environ.get(OUTPUT_DIR_COMPAT_ENV_VAR)
        or default_if_missing
    )
    if not configured:
        return None

    resolved = str(Path(configured).expanduser().resolve())
    # Keep env vars synchronized for cross-server compatibility.
    os.environ[OUTPUT_DIR_PRIMARY_ENV_VAR] = resolved
    os.environ["EXCEL_FILES_PATH"] = resolved
    return resolved

def ensure_excel_extension(file_path: str) -> str:
    """
    Normalize filename and ensure .xlsx extension.

    - trims whitespace and wrapping quotes
    - defaults empty to workbook.xlsx
    - enforces .xlsx extension (case-insensitive)
    - resolves basename-only paths to output dir when configured
    """
    normalized = _clean_user_path(file_path)
    if not normalized:
        normalized = "workbook.xlsx"

    if not normalized.lower().endswith(".xlsx"):
        normalized = f"{normalized}.xlsx"

    files_root = get_files_root()
    if files_root and _is_basename_only(normalized):
        return str(files_root / normalized)
    return normalized

def _checked_paths_error(file_ref: str, checked_paths: list[Path]) -> FileNotFoundError:
    checked = ", ".join(str(path) for path in checked_paths)
    return FileNotFoundError(f"Excel file '{file_ref}' not found. Checked: {checked}")

def resolve_existing_excel_path(file_ref: str) -> Path:
    """Resolve an existing workbook path with basename/case-insensitive fallbacks."""
    cleaned = _clean_user_path(file_ref)
    if not cleaned:
        raise FileNotFoundError("Excel file path cannot be empty.")

    if not cleaned.lower().endswith(".xlsx"):
        cleaned = f"{cleaned}.xlsx"

    requested_path = Path(cleaned)
    checked_paths: list[Path] = []

    # 1) exact path (absolute or cwd-relative)
    exact_candidate = requested_path if requested_path.is_absolute() else requested_path.resolve()
    checked_paths.append(exact_candidate)
    if exact_candidate.exists() and exact_candidate.is_file():
        return exact_candidate.resolve()

    files_root = get_files_root()
    if files_root:
        # 2) output-dir fallback for relative paths
        if not requested_path.is_absolute():
            output_candidate = (files_root / requested_path).resolve()
            checked_paths.append(output_candidate)
            if output_candidate.exists() and output_candidate.is_file():
                return output_candidate

        # 3) basename fallback in output dir
        basename = requested_path.name
        basename_candidate = (files_root / basename).resolve()
        checked_paths.append(basename_candidate)
        if basename_candidate.exists() and basename_candidate.is_file():
            return basename_candidate

        # 4) case-insensitive basename fallback
        lower_basename = basename.lower()
        for candidate in files_root.rglob("*"):
            if candidate.is_file() and candidate.name.lower() == lower_basename:
                return candidate.resolve()

    raise _checked_paths_error(file_ref, checked_paths)

def _validate_writable_target(target_path: Path) -> None:
    """Validate target write access before save/copy operations."""
    if target_path.exists():
        if not os.access(target_path, os.W_OK):
            raise PermissionError(f"Target file is not writable: {target_path}")
        return

    parent = target_path.parent
    if not parent.exists():
        parent.mkdir(parents=True, exist_ok=True)
    if not os.access(parent, os.W_OK):
        raise PermissionError(f"Target directory is not writable: {parent}")

def resolve_target_excel_path(target_ref: str) -> Path:
    """Resolve target save path, using output dir when given basename only."""
    normalized = ensure_excel_extension(target_ref)
    target_path = Path(normalized)
    if not target_path.is_absolute():
        target_path = target_path.resolve()

    _validate_writable_target(target_path)
    return target_path

def get_download_base_url() -> Optional[str]:
    base_url = (
        os.environ.get(DOWNLOAD_BASE_URL_PRIMARY_ENV_VAR)
        or os.environ.get(DOWNLOAD_BASE_URL_COMPAT_ENV_VAR)
        or os.environ.get("EXCEL_DOWNLOAD_BASE_URL")
    )
    if not base_url:
        return None
    return base_url.strip().rstrip("/")

def get_download_signing_secret() -> Optional[str]:
    secret = (
        os.environ.get(DOWNLOAD_SIGNING_SECRET_PRIMARY_ENV_VAR)
        or os.environ.get(DOWNLOAD_SIGNING_SECRET_COMPAT_ENV_VAR)
        or os.environ.get("EXCEL_DOWNLOAD_SIGNING_SECRET")
    )
    if not secret:
        return None
    return secret.strip()

def get_download_url_ttl_seconds() -> int:
    raw_value = (
        os.environ.get(DOWNLOAD_URL_TTL_PRIMARY_ENV_VAR)
        or os.environ.get(DOWNLOAD_URL_TTL_COMPAT_ENV_VAR)
        or os.environ.get("EXCEL_DOWNLOAD_URL_TTL_SECONDS")
        or "300"
    )
    try:
        ttl = int(raw_value)
    except (TypeError, ValueError):
        return 300
    return max(30, ttl)

def build_download_signature(filename: str, expires_at: int, secret: str) -> str:
    payload = f"{filename}:{expires_at}".encode("utf-8")
    return hmac.new(secret.encode("utf-8"), payload, hashlib.sha256).hexdigest()

def _extract_download_filename_from_path(path_value: str) -> Optional[str]:
    prefix = "/files/"
    if not path_value.startswith(prefix):
        return None
    raw_filename = path_value[len(prefix):]
    if not raw_filename:
        return None

    filename = _clean_user_path(unquote(raw_filename))
    if not filename:
        return None
    if Path(filename).name != filename:
        return None
    if not filename.lower().endswith(".xlsx"):
        return None
    return filename

def _client_prefers_html(request: Request) -> bool:
    """Return True when client explicitly accepts HTML responses."""
    accept = (request.headers.get("accept") or "").lower()
    return "text/html" in accept

def _signed_link_error_response(request: Request, signed_status: str) -> Response:
    if _client_prefers_html(request):
        if signed_status == "expired":
            title = "Download Link Expired"
            message = "This link has expired. Please request a new download link."
            hint = "The previous URL is no longer valid for security reasons."
        else:
            title = "Invalid Download Link"
            message = "This download link is invalid. Please request a new link."
            hint = "Please check the full link or generate a fresh signed download URL."

        html = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{title}</title>
  <style>
    :root {{
      --brand-primary: #002554;
      --brand-primary-hover: #0e3a6f;
      --brand-aqua: #2cb1bc;
      --surface-app: #f7f8fa;
      --surface-raised: #ffffff;
      --surface-muted: #f1f3f6;
      --text-primary: #1a1f2b;
      --text-secondary: #4b5563;
      --border-default: #e2e8f0;
      --border-strong: #cbd5e1;
    }}
    @media (prefers-color-scheme: dark) {{
      :root {{
        --surface-app: #0b1526;
        --surface-raised: #0f1b33;
        --surface-muted: #152341;
        --text-primary: #e6e8eb;
        --text-secondary: #c7cdd8;
        --border-default: #1f2a37;
        --border-strong: #2a3646;
      }}
    }}
    * {{
      box-sizing: border-box;
    }}
    body {{
      margin: 0;
      font-family: "Gothce", Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background:
        radial-gradient(1100px 420px at 90% -100px, rgba(44, 177, 188, 0.16), transparent 65%),
        radial-gradient(980px 360px at 10% -120px, rgba(0, 37, 84, 0.18), transparent 60%),
        var(--surface-app);
      color: var(--text-primary);
      display: grid;
      min-height: 100vh;
      place-items: center;
      padding: 22px;
    }}
    .card {{
      max-width: 560px;
      width: 100%;
      background: var(--surface-raised);
      border: 1px solid var(--border-default);
      border-radius: 14px;
      overflow: hidden;
      box-shadow: 0 18px 46px rgba(7, 14, 30, 0.2);
    }}
    .brand {{
      background: linear-gradient(135deg, var(--brand-primary) 0%, var(--brand-primary-hover) 100%);
      color: #ffffff;
      padding: 16px 22px;
      border-bottom: 1px solid rgba(255, 255, 255, 0.2);
    }}
    .wordmark {{
      margin: 0;
      font-size: 0.92rem;
      line-height: 1.2;
      letter-spacing: 0.07em;
      text-transform: uppercase;
      font-weight: 600;
      opacity: 0.95;
    }}
    .product {{
      margin: 4px 0 0;
      font-size: 0.95rem;
      letter-spacing: 0.04em;
      opacity: 0.96;
    }}
    .content {{
      padding: 24px 22px 22px;
    }}
    .status {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      background: var(--surface-muted);
      border: 1px solid var(--border-strong);
      color: var(--text-secondary);
      border-radius: 999px;
      padding: 6px 10px;
      font-size: 0.78rem;
      letter-spacing: 0.03em;
      text-transform: uppercase;
      font-weight: 600;
    }}
    .dot {{
      width: 8px;
      height: 8px;
      border-radius: 50%;
      background: var(--brand-aqua);
      box-shadow: 0 0 0 3px rgba(44, 177, 188, 0.22);
    }}
    h1 {{
      margin: 14px 0 10px;
      font-size: 1.4rem;
      line-height: 1.2;
      color: var(--text-primary);
    }}
    p {{
      margin: 0 0 10px;
      line-height: 1.55;
      color: var(--text-secondary);
      font-size: 0.98rem;
    }}
    .footer {{
      margin-top: 16px;
      padding-top: 14px;
      border-top: 1px solid var(--border-default);
      color: var(--text-secondary);
      font-size: 0.82rem;
      letter-spacing: 0.02em;
    }}
  </style>
</head>
<body>
  <main class="card">
    <header class="brand">
      <p class="wordmark">MATTONI 1873</p>
      <p class="product">DIGITAL | Mchat</p>
    </header>
    <section class="content">
      <span class="status"><span class="dot"></span> Secure Download</span>
      <h1>{title}</h1>
      <p>{message}</p>
      <p>{hint}</p>
      <p class="footer">Mattoni 1873 - M chat</p>
    </section>
  </main>
</body>
</html>"""
        return HTMLResponse(content=html, status_code=403)

    return JSONResponse(
        {"error": "Invalid or expired download signature."},
        status_code=403,
    )

def evaluate_signed_download_request(request: Request) -> str:
    """
    Validate short-lived signed file download URLs.

    Returns:
      - "valid" if request contains a valid signature
      - "expired" if signature is valid but expired
      - "invalid" if signature params are present but invalid
      - "not_attempted" if no signature params are present
    """
    secret = get_download_signing_secret()
    if not secret:
        return "not_attempted"

    exp_raw = request.query_params.get("exp") or request.query_params.get("expires")
    sig = request.query_params.get("sig") or request.query_params.get("signature")
    has_signature_params = any(
        request.query_params.get(key) is not None
        for key in ("exp", "expires", "sig", "signature")
    )
    if not has_signature_params:
        return "not_attempted"

    filename = _extract_download_filename_from_path(request.url.path)
    if not filename:
        return "invalid"

    try:
        expires_at = int(exp_raw or "")
    except (TypeError, ValueError):
        return "invalid"

    if expires_at < int(time.time()):
        return "expired"

    expected = build_download_signature(filename, expires_at, secret)
    if not sig or not secrets.compare_digest(sig, expected):
        return "invalid"

    return "valid"

def build_download_url(file_path: Path) -> Optional[str]:
    """Build download URL when a base URL is configured."""
    base_url = get_download_base_url()
    if not base_url:
        return None

    url = f"{base_url}/files/{quote(file_path.name)}"
    secret = get_download_signing_secret()
    if not secret:
        return url

    expires_at = int(time.time()) + get_download_url_ttl_seconds()
    signature = build_download_signature(file_path.name, expires_at, secret)
    return f"{url}?exp={expires_at}&sig={signature}"

def get_excel_path(filename: str, must_exist: bool = True) -> str:
    """
    Get normalized workbook path.

    For existing files, applies exact/output-dir/case-insensitive fallback logic.
    For new files, resolves basename-only targets into output dir.
    """
    if must_exist:
        return str(resolve_existing_excel_path(filename))
    return str(resolve_target_excel_path(filename))

def get_files_root() -> Optional[Path]:
    """Return the resolved files root path for HTTP file routes."""
    if EXCEL_FILES_PATH is None:
        return None
    return Path(EXCEL_FILES_PATH).resolve()

def resolve_download_path(raw_file_path: str) -> Path:
    """
    Resolve basename-only .xlsx download path under EXCEL_FILES_PATH.

    Invalid names intentionally map to a 404 response.
    """
    files_root = get_files_root()
    if files_root is None:
        raise FileNotFoundError("File not found.")

    filename = _clean_user_path(unquote(raw_file_path))
    if not filename:
        raise FileNotFoundError("File not found.")
    if Path(filename).name != filename:
        raise FileNotFoundError("File not found.")
    if not filename.lower().endswith(".xlsx"):
        raise FileNotFoundError("File not found.")

    requested_path = (files_root / filename).resolve()
    if requested_path.parent != files_root:
        raise FileNotFoundError("File not found.")
    return requested_path

@mcp.custom_route("/files", methods=["GET"])
async def list_generated_files(_: Request) -> Response:
    """
    List files currently available under EXCEL_FILES_PATH.

    This route is available in SSE and streamable HTTP transports.
    """
    files_root = get_files_root()
    if files_root is None:
        return JSONResponse(
            {"error": "EXCEL_FILES_PATH is not configured for this transport."},
            status_code=400,
        )

    if not files_root.exists():
        return JSONResponse({"files_root": str(files_root), "count": 0, "files": []})

    files = []
    for path in sorted(files_root.glob("*.xlsx")):
        if path.is_file() and path.suffix.lower() == ".xlsx":
            rel_path = path.name
            item = {
                "path": rel_path,
                "name": path.name,
                "size_bytes": path.stat().st_size,
            }
            download_url = build_download_url(path)
            if download_url:
                item["download_url"] = download_url
            files.append(
                {
                    **item
                }
            )

    return JSONResponse({"files_root": str(files_root), "count": len(files), "files": files})

@mcp.custom_route("/files/{file_path:path}", methods=["GET"])
async def download_generated_file(request: Request) -> Response:
    """
    Download a file from EXCEL_FILES_PATH by relative path.

    Example:
      GET /files/report.xlsx
    """
    signed_status = evaluate_signed_download_request(request)
    if signed_status in {"expired", "invalid"}:
        return _signed_link_error_response(request, signed_status)

    raw_file_path = request.path_params.get("file_path", "")
    try:
        file_path = resolve_download_path(raw_file_path)
    except FileNotFoundError:
        return JSONResponse({"error": "File not found."}, status_code=404)

    if not file_path.exists() or not file_path.is_file():
        return JSONResponse({"error": "File not found."}, status_code=404)

    media_type, _ = mimetypes.guess_type(str(file_path))
    return FileResponse(
        path=str(file_path),
        media_type=media_type or "application/octet-stream",
        filename=file_path.name,
    )

@mcp.custom_route("/healthz", methods=["GET"])
async def healthz(_: Request) -> Response:
    """Unauthenticated healthcheck endpoint for load balancers."""
    return JSONResponse({"status": "ok"})

@mcp.tool(
    annotations=ToolAnnotations(
        title="Apply Formula",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def apply_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Apply Excel formula to cell.
    Excel formula will write to cell with verification.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        # First validate the formula
        validation = validate_formula_impl(full_path, sheet_name, cell, formula)
        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"
            
        # If valid, apply the formula
        from excel_mcp.calculations import apply_formula as apply_formula_impl
        result = apply_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Validate Formula Syntax",
        readOnlyHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Validate Excel formula syntax without applying it.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = validate_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Format Range",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None
) -> str:
    """
    Apply formatting to a range of cells.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.formatting import format_range as format_range_func
        
        # Convert None values to appropriate defaults for the underlying function
        format_range_func(
            filepath=full_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,  # This can be None
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,  # This can be None
            font_color=font_color,  # This can be None
            bg_color=bg_color,  # This can be None
            border_style=border_style,  # This can be None
            border_color=border_color,  # This can be None
            number_format=number_format,  # This can be None
            alignment=alignment,  # This can be None
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,  # This can be None
            conditional_format=conditional_format  # This can be None
        )
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Read Data from Excel",
        readOnlyHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def read_data_from_excel(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False
) -> str:
    """
    Read data from Excel worksheet with cell metadata including validation rules.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell (default A1)
        end_cell: Ending cell (optional, auto-expands if not provided)
        preview_only: Whether to return preview only
    
    Returns:  
    JSON string containing structured cell data with validation metadata.
    Each cell includes: address, value, row, column, and validation info (if any).

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.data import read_excel_range_with_metadata
        result = read_excel_range_with_metadata(
            full_path, 
            sheet_name, 
            start_cell, 
            end_cell
        )
        if not result or not result.get("cells"):
            return "No data found in specified range"
            
        # Return as formatted JSON string
        return json.dumps(result, indent=2, default=str)
        
    except Exception as e:
        logger.error(f"Error reading data: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="List Excel Files",
        readOnlyHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def list_excel_files(directory: str = "") -> str:
    """
    List .xlsx files in a directory.

    If directory is empty or ".", and an output dir is configured, lists output dir.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        files_root = get_files_root()
        requested_dir = _clean_user_path(directory)

        if (requested_dir == "" or requested_dir == ".") and files_root:
            target_dir = files_root
        elif requested_dir == "" or requested_dir == ".":
            target_dir = Path(".").resolve()
        else:
            requested_path = Path(requested_dir)
            if requested_path.is_absolute():
                target_dir = requested_path.resolve()
            elif files_root:
                target_dir = (files_root / requested_path).resolve()
            else:
                target_dir = requested_path.resolve()

        if not target_dir.exists() or not target_dir.is_dir():
            return f"Error: Directory not found: {target_dir}"

        files: list[dict[str, Any]] = []
        for candidate in sorted(target_dir.glob("*.xlsx")):
            if not candidate.is_file():
                continue
            item: dict[str, Any] = {
                "name": candidate.name,
                "path": str(candidate.resolve()),
                "size_bytes": candidate.stat().st_size,
            }
            download_url = build_download_url(candidate)
            if download_url:
                item["download_url"] = download_url
            files.append(item)

        payload = {
            "directory": str(target_dir),
            "count": len(files),
            "files": files,
        }
        return json.dumps(payload, indent=2)
    except Exception as e:
        logger.error(f"Error listing Excel files: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Write Data to Excel",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def write_data_to_excel(
    filepath: str,
    sheet_name: str,
    data: List[List],
    start_cell: str = "A1",
) -> str:
    """
    Write data to Excel worksheet.
    Excel formula will write to cell without any verification.

    PARAMETERS:  
    filepath: Path to Excel file
    sheet_name: Name of worksheet to write to
    data: List of lists containing data to write to the worksheet, sublists are assumed to be rows
    start_cell: Cell to start writing to, default is "A1"

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = write_data(full_path, sheet_name, data, start_cell)
        return result["message"]
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Create Workbook",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def create_workbook(filepath: str) -> str:
    """
    Create new Excel workbook.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath, must_exist=False)
        from excel_mcp.workbook import create_workbook as create_workbook_impl
        create_workbook_impl(full_path)
        created_path = Path(full_path).resolve()
        response_payload: Dict[str, Any] = {
            "message": "Workbook created successfully.",
            "file_path": str(created_path),
            "file_size_bytes": created_path.stat().st_size if created_path.exists() else 0,
        }
        download_url = build_download_url(created_path)
        if download_url:
            response_payload["download_url"] = download_url
        return json.dumps(response_payload, indent=2)
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Save Excel File",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def save_excel_file(file_path: str, source_filename: str) -> str:
    """
    Save/copy an existing workbook to a target path.

    The source file is resolved using robust lookup:
    exact path, output-dir fallback, and case-insensitive basename fallback.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        source_path = resolve_existing_excel_path(source_filename)
        target_path = resolve_target_excel_path(file_path)

        if source_path.resolve() == target_path.resolve():
            payload: Dict[str, Any] = {
                "message": "No-op: source and target paths are the same.",
                "file_path": str(target_path),
                "file_size_bytes": target_path.stat().st_size if target_path.exists() else 0,
            }
            download_url = build_download_url(target_path)
            if download_url:
                payload["download_url"] = download_url
            return json.dumps(payload, indent=2)

        shutil.copy2(source_path, target_path)
        if not target_path.exists():
            raise WorkbookError(f"Saved file was not found after copy: {target_path}")

        payload = {
            "message": "Workbook saved successfully.",
            "file_path": str(target_path),
            "file_size_bytes": target_path.stat().st_size,
        }
        download_url = build_download_url(target_path)
        if download_url:
            payload["download_url"] = download_url
        return json.dumps(payload, indent=2)
    except (FileNotFoundError, PermissionError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error saving workbook: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Create Worksheet",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def create_worksheet(filepath: str, sheet_name: str) -> str:
    """
    Create new worksheet in workbook.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_sheet as create_worksheet_impl
        result = create_worksheet_impl(full_path, sheet_name)
        return result["message"]
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Create Chart",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def create_chart(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = ""
) -> str:
    """
    Create chart in worksheet.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = create_chart_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            chart_type=chart_type,
            target_cell=target_cell,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis
        )
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Create Pivot Table",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "mean"
) -> str:
    """
    Create pivot table in worksheet.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = create_pivot_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            rows=rows,
            values=values,
            columns=columns or [],
            agg_func=agg_func
        )
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Create Table",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def create_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: Optional[str] = None,
    table_style: str = "TableStyleMedium9"
) -> str:
    """
    Creates a native Excel table from a specified range of data.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = create_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            table_name=table_name,
            table_style=table_style
        )
        return result["message"]
    except DataError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Copy Worksheet",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def copy_worksheet(
    filepath: str,
    source_sheet: str,
    target_sheet: str
) -> str:
    """
    Copy worksheet within workbook.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = copy_sheet(full_path, source_sheet, target_sheet)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Delete Worksheet",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def delete_worksheet(
    filepath: str,
    sheet_name: str
) -> str:
    """
    Delete worksheet from workbook.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = delete_sheet(full_path, sheet_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Rename Worksheet",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def rename_worksheet(
    filepath: str,
    old_name: str,
    new_name: str
) -> str:
    """
    Rename worksheet in workbook.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = rename_sheet(full_path, old_name, new_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Get Workbook Metadata",
        readOnlyHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def get_workbook_metadata(
    filepath: str,
    include_ranges: bool = False
) -> str:
    """
    Get metadata about workbook including sheets, ranges, etc.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = get_workbook_info(full_path, include_ranges=include_ranges)
        return str(result)
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Merge Cells",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def merge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """
    Merge a range of cells.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = merge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Unmerge Cells",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def unmerge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """
    Unmerge a range of cells.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = unmerge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Get Merged Cells",
        readOnlyHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def get_merged_cells(filepath: str, sheet_name: str) -> str:
    """
    Get merged cells in a worksheet.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        return str(get_merged_ranges(full_path, sheet_name))
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting merged cells: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Copy Range",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def copy_range(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None
) -> str:
    """
    Copy a range of cells to another location.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import copy_range_operation
        result = copy_range_operation(
            full_path,
            sheet_name,
            source_start,
            source_end,
            target_start,
            target_sheet or sheet_name  # Use source sheet if target_sheet is None
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Delete Range",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def delete_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    shift_direction: str = "up"
) -> str:
    """
    Delete a range of cells and shift remaining cells.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import delete_range_operation
        result = delete_range_operation(
            full_path,
            sheet_name,
            start_cell,
            end_cell,
            shift_direction
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Validate Excel Range",
        readOnlyHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def validate_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None
) -> str:
    """
    Validate if a range exists and is properly formatted.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = validate_range_impl(full_path, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Get Data Validation Info",
        readOnlyHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def get_data_validation_info(
    filepath: str,
    sheet_name: str
) -> str:
    """
    Get all data validation rules in a worksheet.
    
    This tool helps identify which cell ranges have validation rules
    and what types of validation are applied.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        
    Returns:
        JSON string containing all validation rules in the worksheet

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        from excel_mcp.cell_validation import get_all_validation_ranges
        
        wb = load_workbook(full_path, read_only=False)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
            
        ws = wb[sheet_name]
        validations = get_all_validation_ranges(ws)
        wb.close()
        
        if not validations:
            return "No data validation rules found in this worksheet"
            
        import json
        return json.dumps({
            "sheet_name": sheet_name,
            "validation_rules": validations
        }, indent=2, default=str)
        
    except Exception as e:
        logger.error(f"Error getting validation info: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Insert Rows",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def insert_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    count: int = 1
) -> str:
    """
    Insert one or more rows starting at the specified row.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = insert_row(full_path, sheet_name, start_row, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting rows: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Insert Columns",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def insert_columns(
    filepath: str,
    sheet_name: str,
    start_col: int,
    count: int = 1
) -> str:
    """
    Insert one or more columns starting at the specified column.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = insert_cols(full_path, sheet_name, start_col, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting columns: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Delete Rows",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def delete_sheet_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    count: int = 1
) -> str:
    """
    Delete one or more rows starting at the specified row.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = delete_rows(full_path, sheet_name, start_row, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting rows: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Delete Columns",
        destructiveHint=True,
    ),
)
@append_finisher_hint_to_tool_output
def delete_sheet_columns(
    filepath: str,
    sheet_name: str,
    start_col: int,
    count: int = 1
) -> str:
    """
    Delete one or more columns starting at the specified column.

    Finisher hint: when the user request appears complete, call
        save_document_mcp_word-mcp so the user receives a download link.
    """
    try:
        full_path = get_excel_path(filepath)
        result = delete_cols(full_path, sheet_name, start_col, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting columns: {e}")
        raise

def run_sse():
    """Run Excel MCP server in SSE mode."""
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = resolve_output_dir(default_if_missing="./excel_files")
    if EXCEL_FILES_PATH is None:
        raise RuntimeError("Failed to resolve output directory for SSE transport.")
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with SSE transport (files directory: {EXCEL_FILES_PATH})")
        mcp.run(transport="sse")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_streamable_http():
    """Run Excel MCP server in streamable HTTP mode."""
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = resolve_output_dir(default_if_missing="./excel_files")
    if EXCEL_FILES_PATH is None:
        raise RuntimeError("Failed to resolve output directory for streamable HTTP transport.")
    api_key = os.environ.get(API_KEY_ENV_VAR, "").strip()
    api_key_header = os.environ.get(API_KEY_HEADER_ENV_VAR, "x-api-key").strip() or "x-api-key"
    host = os.environ.get("FASTMCP_HOST", "0.0.0.0")
    port = int(os.environ.get("FASTMCP_PORT", "8017"))
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with streamable HTTP transport (files directory: {EXCEL_FILES_PATH})")
        if api_key:
            # When API key auth is enabled, run the ASGI app with middleware
            # so both /mcp and custom routes (/files, /healthz) are protected.
            import uvicorn

            app = mcp.streamable_http_app()
            app.add_middleware(
                APIKeyMiddleware,
                api_key=api_key,
                header_name=api_key_header,
                exempt_paths=["/healthz"],
            )
            logger.info(f"API key auth enabled for HTTP endpoints via header '{api_key_header}'")
            uvicorn.run(app, host=host, port=port)
        else:
            mcp.run(transport="streamable-http")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_stdio():
    """Run Excel MCP server in stdio mode."""
    global EXCEL_FILES_PATH
    # Keep stdio backward-compatible: only use output dir if explicitly configured.
    EXCEL_FILES_PATH = resolve_output_dir(default_if_missing=None)
    if EXCEL_FILES_PATH:
        os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info("Starting Excel MCP server with stdio transport")
        mcp.run(transport="stdio")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")
